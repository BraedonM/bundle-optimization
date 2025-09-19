from PyQt6.QtWidgets import QMessageBox, QWidget, QApplication
from BundleQtGui import Ui_BundleOptimizer
from PyQt6 import QtGui
from PyQt6 import QtCore

import openpyxl
import pandas as pd
from tkinter import filedialog
import os
from math import ceil, floor
from datetime import datetime
import time
import numpy as np
import warnings
import ctypes
import cProfile

from bundle_classes import SKU, create_packaging_classes
from bundle_visualize import visualize_bundles
from bundle_packing import pack_skus

def excepthook(type, value, traceback):
    """
    Handle exceptions
    """
    try:
        errorString = (
            f"An error occurred:\n\n{value}"
            + f"\n\n{traceback.tb_frame.f_code.co_name} in "
            + f"{(traceback.tb_frame.f_code.co_filename).split("\\")[-1]} "
            + f"(line {traceback.tb_lineno})"
        )
    except Exception:
        errorString = "An error occurred."

    QMessageBox.critical(None, "Error", errorString)

class ProgramGUI:
    def __init__(self):
        self.Widget = QWidget()
        self.ui = Ui_BundleOptimizer()
        self.ui.setupUi(self.Widget)

        self.Widget.show()
        self.setupUi()

    def setupUi(self):
        # Set icon
        myappid = 'com.aionex.bundleoptimizer'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

        icon_path = os.path.join(os.path.dirname(__file__), 'app_icon_alt.ico')
        self.Widget.setWindowIcon(QtGui.QIcon(icon_path))
        warnings.simplefilter(action='ignore', category=FutureWarning)

        self.ui.helpButton.setIcon(QtGui.QIcon(os.path.join(os.path.dirname(__file__), 'help_icon.png')))

        # Initialize values and connect signals (camelCase, while helper methods are snake_case)
        self.ui.fileBrowse.clicked.connect(self.getInputWorkbook)
        self.ui.appendBrowse.clicked.connect(self.getAppendWorkbook)
        self.ui.openExample.clicked.connect(self.openExampleFile)
        self.ui.optimizeBundles.clicked.connect(self.optimizeBundles)
        self.ui.openImages.clicked.connect(self.openImages)
        self.ui.openExcel.clicked.connect(self.openExcel)
        self.ui.helpButton.clicked.connect(self.openHelp)
        self.ui.metricButton.clicked.connect(self.setMetricUnits)
        self.ui.imperialButton.clicked.connect(self.setImperialUnits)

        self.workingDir = None  # to hold the directory of the selected Excel file
        self.disabledButton = "background-color: rgb(39, 39, 39); color: rgb(255, 255, 255);"
        self.enabledButton = "background-color: rgb(0, 90, 180); color: rgb(255, 255, 255);"
        self.setMetricUnits()  # Set default units to metric
        self.set_unit = 'metric'  # to hold the current unit system

## QT Connection Methods (camelCase)

    def setMetricUnits(self):
        """
        Set the units to metric (mm, kg)
        """
        self.unit = 'metric'
        self.ui.metricButton.setStyleSheet(self.enabledButton)
        self.ui.imperialButton.setStyleSheet(self.disabledButton)

    def setImperialUnits(self):
        """
        Set the units to imperial (inches, lbs)
        """
        self.unit = 'imperial'
        self.ui.imperialButton.setStyleSheet(self.enabledButton)
        self.ui.metricButton.setStyleSheet(self.disabledButton)

    def getInputWorkbook(self):
        """
        Open a dialog to pick an Excel file, return the workbook object
        """
        # get excel file path from user
        path = filedialog.askopenfilename(
            title="Select an Excel file",
            filetypes=[("Excel files", "*.xlsx; *.xls; *.xlsm")],
        )
        if not path:
            return

        self.workingDir = path.rsplit('/', 1)[0]  # get the directory of the selected file

        # load the workbook and read the sheet "SO_Input"
        self.ui.excelDir.setText(path)

    def getAppendWorkbook(self):
        """
        Get the workbook of Optimized data to append new data to
        """
        path = filedialog.askopenfilename(
            title="Select an Excel file to append optimized data to",
            filetypes=[("Excel files", "*.xlsx; *.xls; *.xlsm")],
        )
        if not path:
            return

        # load the workbook and read the sheet "SO_Input"
        self.ui.appendDir.setText(path)

    def openExampleFile(self):
        """
        Open the example Excel file to the user, in Excel
        """
        example_path = os.path.join(os.path.dirname(__file__), 'SO_Input_Example.xlsx')
        if os.path.exists(example_path):
            os.startfile(example_path)
        else:
            self.show_alert("File Not Found", "Example file not found.")

    def optimizeBundles(self):
        """
        Optimize bundles based on the data from the Excel file
        """
        self.maxWidth = 508
        self.maxHeight = 508
        self.maxLength = 3680
        self.missingDataSKUs = []  # to hold SKUs that are missing data in the Excel file
        self.removed_skus = []  # to hold SKUs that were removed during optimization
        self.mach1_skus = []  # to hold SKUs that are packed with Mach1
        self.append_data = False  # to indicate if we are appending data to an existing workbook
        self.set_unit = self.unit

        self.ui.progressLabel.setText("Getting data for new orders...")
        self.ui.progressBar.setValue(10)
        # Get data from input workbook
        try:
            self.workbook = openpyxl.load_workbook(self.ui.excelDir.text(), data_only=True)
            data = self.get_data(self.workbook)
        except Exception as e:
            self.show_alert("Error", "Unable to retrieve data from the Excel file.\nEnsure the file is not open and the path is correct.", "error")
            self.ui.progressBar.setValue(0)
            self.ui.progressLabel.setText("")
            return

        # Get packaging and filler data from the packaging_data file
        try:
            packaging_data = self.get_packaging_data()
            self.packaging_height, self.packaging_width, self.lumber_height = create_packaging_classes(packaging_data)
        except Exception as e:
            self.show_alert("Error", f"Unable to retrieve data from the packaging data file. Error: {e}", "error")
            self.ui.progressBar.setValue(0)
            self.ui.progressLabel.setText("")
            return

        # get unique orders
        unique_orders = list(data['OrderNbr'].unique())

        if self.ui.appendDir.text():
            self.ui.progressLabel.setText("Getting data from existing bundles...")
            self.ui.progressBar.setValue(15)
            QApplication.processEvents()
            self.appendWorkbook = openpyxl.load_workbook(self.ui.appendDir.text())
            self.append_data = True
            unique_orders, self.appendWorkbook = self.remove_optimized_orders(unique_orders, self.appendWorkbook)
            if not unique_orders:
                self.show_alert("No Orders", "All selected orders have already been optimized in the append workbook.", "info")
                self.ui.progressBar.setValue(0)
                self.ui.progressLabel.setText("")
                return
            elif unique_orders == -1:
                self.show_alert("Error", "Unit mismatch between the program input and the append workbook.\nPlease ensure the units of measurement are consistent.", "error")
                self.ui.progressBar.setValue(0)
                self.ui.progressLabel.setText("")
                return
        else:
            # delete existing file
            self.append_data = False
            try:
                if os.path.exists(f"{self.workingDir}/Optimized_Bundles.xlsx"):
                    os.remove(f"{self.workingDir}/Optimized_Bundles.xlsx")
            except Exception as e:
                self.show_alert("Error", f"Error accessing Optimized_Bundles.xlsx file: {e}", "error")
                self.ui.progressBar.setValue(0)
                self.ui.progressLabel.setText("")
                return

        # get array of rows in sorted_data that belong to each order
        order_rows = {order: data[data['OrderNbr'] == order] for order in unique_orders}

        # create SKU objects for each order
        order_skus = self.create_sku_objects(order_rows)

        order_skus = self.remove_invalids(order_skus)

        # convert orders from numpy floats to ints
        for order in reversed(order_skus.keys()):
            if isinstance(order, np.float64) or isinstance(order, np.int64):
                order_skus[int(order)] = order_skus.pop(order)

        self.ui.progressBar.setValue(20)
        images_dir = f"{self.workingDir}/images"
        os.makedirs(images_dir, exist_ok=True)

        images_dir = f"{self.workingDir}/images"
        os.makedirs(images_dir, exist_ok=True)

        # pack each order's SKUs into bundles
        order_bundles = {}
        for order, skus in order_skus.items():
            self.ui.progressLabel.setText(f"Packing order {str(order).split('.')[0]}...")
            # pause to update the GUI
            QApplication.processEvents()
            if skus == []:
                order_bundles[order] = []
                continue
            self.ui.progressBar.setValue(int(round(20 + 70 * (list(order_skus.keys()).index(order)) / len(order_skus))))
            bundles, self.removed_skus = pack_skus(skus, self.maxWidth, self.maxHeight, self.mach1_skus)
            if bundles == -1:
                self.show_alert("Error", "Cannot mix MACH1 and MACH5 SKUs in the same bundle override.", "error")
                self.ui.progressBar.setValue(0)
                self.ui.progressLabel.setText("")
                return
            order_bundles[order] = bundles

            visualize_bundles(bundles, f"{images_dir}/Order_{order}.png", self.set_unit, self.packaging_height, self.packaging_width, self.lumber_height)
        self.images_dir = images_dir

        # write the packed bundles to a new sheet in the workbook
        if self.append_data:
            self.workbook = self.appendWorkbook
        self.ui.progressBar.setValue(90)
        self.ui.progressLabel.setText("Writing optimized bundles to Excel...")
        QApplication.processEvents()
        # self.missingDataSKUs.extend(self.removed_skus)  # add removed SKUs to missing data SKUs
        self.write_optimized_bundles(self.workbook, order_bundles)

        self.ui.progressBar.setValue(100)
        self.ui.progressLabel.setText("Packing complete!")

        if self.missingDataSKUs:
            self.show_alert("Missing Data", "There exist InventoryIDs that are missing data in the Excel file\nand have been excluded from optimization.\n\nThey can be found under bundle \'0\' for each order in the optimization file.", "warning")

    def openImages(self):
        """
        Open the images directory in the file explorer
        """
        if hasattr(self, 'images_dir'):
            os.startfile(self.images_dir)
        else:
            self.show_alert("No Images", "No images have been generated yet. Please optimize bundles first.", "error")

    def openExcel(self):
        """
        Open the optimized bundles Excel file in the file explorer
        """
        if not self.workingDir:
            self.show_alert("Error", "No file found", "error")
            return
        excel_path = f"{self.workingDir}/Optimized_Bundles.xlsx"
        if os.path.exists(excel_path):
            os.startfile(excel_path)
        else:
            self.show_alert("Error", "Optimized bundles file not found. Please optimize bundles first.", "error")

    def openHelp(self):
        """
        Open the help file in the file explorer
        """
        # def optimize_test():
        #     self.optimizeBundles()
        # cProfile.runctx('self.optimizeBundles()', None, locals())
        self.show_alert("Help", """
        <p>Bundle Optimization Tool v1.0<br>
        For additional support, please contact Aionex Solutions Ltd.<br>
        (604-309-8975)</p>

        <p><b>Program Releases and Updates:</b><br>
        <a href="https://github.com/BraedonM/bundle-optimization/releases">
        https://github.com/BraedonM/bundle-optimization/releases</a></p>

        <p>To use this tool, follow these steps:</p>

        <ol>
            <li>Click on the first 'Browse' button to select your Excel file containing SKU data.<br>
                - If you are unsure how to format your Excel file,<br>
                  you can click on 'Open Example File' to open a sample file.<br>
                - Please format your Excel file according to the example provided.
            </li>

            <li>If you want to append the optimized data to an existing file,<br>
                click on the second 'Browse' button to select the Excel file<br>
                where you want to append the optimized data.
            </li>

            <li>Click on 'Perform Bundle Optimization' to start the optimization process.<br>
                Wait for the progress bar to complete.
            </li>

            <li>Once the optimization is complete, you can:<br>
                - Click on 'Open Images Folder' to view the generated bundle images.<br>
                - Click on 'Open Resultant Excel File' to view the optimized bundle data in an Excel file.
            </li>
        </ol>

        <p><b>NOTE:</b> All files are generated in the same directory as the input Excel file.</p>
        """, type="help")

## Helper Methods (snake_case)

    def get_sub_bundle_data_sheets(self):
        """
        Open a dialog to pick an Excel file, return the sheet "Sub-Bundle_Data"
        """
        # load the workbook and read the sheet "Sub-Bundle_Data"
        path = os.path.join(os.path.dirname(__file__), 'Sub-Bundle_Data.xlsx')
        workbook = openpyxl.load_workbook(path)
        sb_data = workbook["Sub-Bundle_Data"]
        if not sb_data:
            self.show_alert("Warning", "Sheet 'Sub-Bundle_Data' not found in file.")
            return None
        mach1_skus = workbook["MACH1_SKUs"]
        if not mach1_skus:
            self.show_alert("Warning", "Sheet 'MACH1_SKUs' not found in file.")
            return None
        return sb_data, mach1_skus

    def remove_optimized_orders(self, orders, workbook):
        """
        Remove orders that are already optimized in the workbook
        """
        if "Optimized_Bundles" not in workbook.sheetnames:
            return orders, workbook
        
        optimized_sheet = workbook["Optimized_Bundles"]
        # remove table formatting if it exists
        if "OptimizedBundlesTable" in optimized_sheet.tables:
            del optimized_sheet.tables["OptimizedBundlesTable"]

        # Check if units match
        for header in optimized_sheet[1]:
            if 'Width' in header.value or 'Height' in header.value or 'Length' in header.value:
                if self.set_unit == 'imperial' and '_mm' in header.value:
                    return -1, workbook  # units mismatch
                if self.set_unit == 'metric' and '_in' in header.value:
                    return -1, workbook

        # find orders that have been updated since last optimization (the 'OptimizedOn' column is earlier than the 'LastModifiedOn' column)
        orders_to_remove = []
        found_order = False
        reverted_overrides = []
        order_rows = {}
        # get the last modified date of the order from the optimized sheet
        for rowIdx in range(optimized_sheet.max_row, 0, -1):
            row = optimized_sheet[rowIdx]
            if not any(cell.value for cell in row):
                found_order = False
                continue  # skip empty rows
            orderNbr = row[1].value
            if orderNbr in orders:
                # remove the order from the file so it can be re-optimized
                if orderNbr not in order_rows:
                    order_rows[orderNbr] = []
                if not found_order:
                    found_order = True
                    order_rows[orderNbr].append(rowIdx + 1)
                order_rows[orderNbr].append(rowIdx)

                if type(row[2].value) is not int:
                    continue
                last_modified_on = row[-2].value
                optimized_on = row[-1].value
                # convert to datetime if not None
                if last_modified_on and optimized_on:
                    last_modified_on = datetime.strptime(last_modified_on, "%Y-%m-%d")
                    optimized_on = datetime.strptime(optimized_on, "%Y-%m-%d")
                    if last_modified_on <= optimized_on:

                        override = row[3].value
                        sku_id = row[7].value.strip()

                        if override and not str(sku_id).startswith('Pack_'):
                            if (orderNbr not in reverted_overrides
                                and orderNbr not in self.override_orders
                                and orderNbr in orders):
                                reverted_overrides.append(orderNbr)

                        elif orderNbr not in self.override_orders: # modified earlier than it was optimized
                            # don't re-optimize this order
                            orders_to_remove.append(orderNbr)
        # remove the orders that have been optimized
        orders_to_remove = set(orders_to_remove)  # convert to set for faster lookup
        orders_to_remove = [order for order in orders_to_remove if order not in reverted_overrides]
        orders = [order for order in orders if order not in orders_to_remove]

        for order in orders:
            if order in order_rows:
                for rowIdx in order_rows[order]:
                    optimized_sheet.row_dimensions[rowIdx].outlineLevel = 0
                    optimized_sheet.delete_rows(rowIdx, 1)

        # save the workbook after removing orders
        try:
            if optimized_sheet.max_row != 1:
                optimized_sheet.append([])  # add a blank row at the end to separate orders
            workbook.save(self.ui.appendDir.text())
        except Exception as e:
            self.show_alert("Error", f"Error with the existing optimized data file.\nEnsure the path is correct and the file is not open. Error: {e}", "error")
            return [], workbook
        return orders, workbook

    def get_data(self, workbook):
        """
        Read data from the 'SO_Input' sheet of the workbook
        """
        # get the "SO_PackExportData" sheet
        so_input = workbook["SO-PackExportData"]
        if not so_input:
            self.show_alert("Warning", "Sheet 'SO-PackExportData' is empty or not found. Using first sheet in the file instead.")
            so_input = workbook.active  # fallback to the first sheet if is not found
        sb_data, mach1_skus_data = self.get_sub_bundle_data_sheets()

        df = pd.DataFrame(columns=[cell.value for cell in so_input[1]])
        sb_df = pd.DataFrame(columns=[cell.value for cell in sb_data[2]])
        # read all rows from the sheets
        for row in so_input.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            df.loc[len(df)] = row
        for row in sb_data.iter_rows(min_row=3, values_only=True):
            # check for any empty cells in the row
            if any(cell is None for cell in [el for el in row][2:7]):
                continue
            sb_df.loc[len(sb_df)] = row

        # check if the required columns are present
        self.headers = ["OrderType", "OrderNbr", "Bdl_Override", "InventoryID", "Quantity", "Pcs/Bundle", "Can_be_bottom",
                   "Dim_shrink", "Width_mm", "Height_mm", "Length_mm", "Weight_kg", "UOM", "Description",
                   "ShipTo", "AddressLine1", "AddressLine2", "City", "State", "Country", "Status", "OrderDate",
                   "ProdReleaseDate", "SchedShipDate", "TargetArrival", "NotBefore", "ShipVia", "LastModifiedOn"]
        for col in self.headers:
            if col not in df.columns:
                # add the column with default values
                df[col] = None

        # put data from sb_df into df rows
        for _, row in sb_df.iterrows():
            sku_id = row['SKU']
            df_rows = []
            if row['Bottom Row Acceptable']:
                can_be_bottom = True
            else:
                can_be_bottom = False
            for index, df_row in df.iterrows():
                try:
                    if sku_id in df_row['InventoryID']:
                        df_rows.append(index)
                except TypeError:
                    # if df_row['InventoryID'] is None, skip this row
                    continue
            if row['Partial Dim To Reduce'] is None:
                row['Partial Dim To Reduce'] = ''
            for df_row_idx in df_rows:
                # update the Pcs/Bundle column with the value from sb_df
                df.loc[df_row_idx, 'Pcs/Bundle'] = row['Qty/bundle']
                df.loc[df_row_idx, 'Width_mm'] = row['Width (mm)']
                df.loc[df_row_idx, 'Height_mm'] = row['Height (mm)']
                df.loc[df_row_idx, 'Length_mm'] = row['Length (mm)']
                df.loc[df_row_idx, 'Weight_kg'] = float(row['Weight kg/bundle'])# / row['Qty/bundle'] * df.loc[df_row_idx, 'BaseOrderQty'])
                df.loc[df_row_idx, 'Dim_shrink'] = row['Partial Dim To Reduce']
                df.loc[df_row_idx, 'Can_be_bottom'] = can_be_bottom

        # iterate through df and convert qty (in pieces) to qty (in bundles)
        df['Quantity'] = df['BaseOrderQty'].astype(float)
        seen_skus = {}  # to track seen SKUs in case of multiple entries
        self.override_orders = []
        for index, row in df.iterrows():
            if row['Pcs/Bundle'] is not None and row['Quantity'] is not None:
                try:
                    if row['Bdl_Override'] and row['OrderNbr'] not in self.override_orders:
                        self.override_orders.append(row['OrderNbr'])
                    # convert Quantity to Pcs/Bundle
                    whole_qty = floor(abs(row['Quantity']) / row['Pcs/Bundle'])
                    fraction_remaining = (ceil(abs(row['Quantity'])) % row['Pcs/Bundle']) / row['Pcs/Bundle']

                    if f'{row['OrderNbr']}_{row['InventoryID']}_{row['Bdl_Override']}' in seen_skus.keys():
                        # add new qty to existing SKU data
                        df.loc[seen_skus[f'{row['OrderNbr']}_{row['InventoryID']}_{row['Bdl_Override']}'], 'Quantity'] += (float(whole_qty) + float(fraction_remaining))
                        df.loc[index, 'Quantity'] = 0  # set current row to 0
                    else:
                        # create new SKU entry
                        seen_skus[f'{row['OrderNbr']}_{row['InventoryID']}_{row['Bdl_Override']}'] = index
                        df.loc[index, 'Quantity'] = float(whole_qty) + float(fraction_remaining)

                except ZeroDivisionError:
                    self.show_alert("Error", f"Pcs/Bundle cannot be zero for SKU {row['InventoryID']}. Please check the input data.", "error")
                    return pd.DataFrame()

        # Get MACH1 SKU identifiers
        for row in mach1_skus_data.iter_rows(min_row=2, values_only=True):
            sku_id = row[0].strip()
            self.mach1_skus.append(sku_id)

        return df

    def get_packaging_data(self):
        """
        Read data from the 'Packaging_Data' file
        """
        path = os.path.join(os.path.dirname(__file__), 'Packaging_Data.xlsx')
        if not os.path.exists(path):
            raise FileNotFoundError("Packaging_Data.xlsx file not found.")
        workbook = openpyxl.load_workbook(path, data_only=True)
        packaging_data = {}

        sheet = workbook['Packaging_Data']
        if not sheet:
            raise ValueError("Sheet 'Packaging_Data' is empty or not found.")
        df = pd.DataFrame(columns=[cell.value for cell in sheet[1]])
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            df.loc[len(df)] = row
        packaging_data['Packaging_Data'] = df

        data_dict = {}
        # Break data into a dictionary
        for pidIdx, pid in enumerate(df['PID']):
            data_dict[pid] = df.iloc[pidIdx]
        return data_dict

    def remove_invalids(self, order_skus: dict) -> dict:
        """
        Iterate through the order_skus dictionary and remove any SKUs that have None as width, height, length, or weight.
        """
        for order, skus in order_skus.items():
            valid_skus = []
            for sku in skus:
                if None in (sku.width, sku.height, sku.length, sku.weight):
                    if sku.id not in self.missingDataSKUs:
                        self.missingDataSKUs.append(sku)  # add to missing data SKUs list
                else:
                    valid_skus.append(sku)
            order_skus[order] = valid_skus
        return order_skus

    def create_sku_objects(self, order_rows: dict):
        """
        Create arrays of SKU objects for each order
        """
        order_skus = {}
        for order, rows in order_rows.items():
            skus = []
            for _, row in rows.iterrows():
                # get quantity of SKU from the row
                quantity = row['Quantity']
                invID = row['InventoryID'].strip()
                newLength = row['Length_mm']
                if newLength is not None:
                    if 3600 <= newLength <= 3700:
                        newLength = 3650

                if type(quantity) is float:
                    # partial sub-bundle
                    remainder = quantity - floor(quantity)
                    if remainder > 0:
                        width, height = self.shrink_to_square(row['Width_mm'], row['Height_mm'], remainder, row['Dim_shrink'])
                        new_invID = f"{invID}_Partial"
                        sku = SKU(
                            id=new_invID,
                            bundleqty=row['Pcs/Bundle'] * remainder,
                            width=width,
                            height=height,
                            length=newLength,
                            weight=row['Weight_kg'] * remainder,
                            desc=row['Description'],
                            can_be_bottom=row['Can_be_bottom'],
                            data={
                                'OrderType': row['OrderType'],
                                'OrderNbr': row['OrderNbr'],
                                'UOM': row['UOM'],
                                'Bdl_Override': row['Bdl_Override'] if pd.notna(row['Bdl_Override']) else None,
                                'ShipTo': row['ShipTo'],
                                'AddressLine1': row['AddressLine1'],
                                'AddressLine2': row['AddressLine2'],
                                'City': row['City'],
                                'State': row['State'],
                                'Country': row['Country'],
                                'Status': row['Status'],
                                'OrderDate': row['OrderDate'],
                                'ProdReleaseDate': row['ProdReleaseDate'],
                                'SchedShipDate': row['SchedShipDate'],
                                'TargetArrival': row['TargetArrival'],
                                'NotBefore': row['NotBefore'],
                                'ShipVia': row['ShipVia'],
                                'LastModifiedOn': row['LastModifiedOn']
                            }
                        )
                        skus.append(sku)
                    quantity = floor(quantity)  # convert to whole number for the rest of the SKUs

                for _ in range(int(abs(ceil(quantity)))):
                    sku = SKU(
                        id=invID,
                        bundleqty=row['Pcs/Bundle'],
                        width=row['Width_mm'],
                        height=row['Height_mm'],
                        length=newLength,
                        weight=row['Weight_kg'],
                        desc=row['Description'],
                        can_be_bottom=row['Can_be_bottom'],
                        data={
                            'OrderType': row['OrderType'],
                            'OrderNbr': row['OrderNbr'],
                            'UOM': row['UOM'],
                            'Bdl_Override': row['Bdl_Override'] if pd.notna(row['Bdl_Override']) else None,
                            'ShipTo': row['ShipTo'],
                            'AddressLine1': row['AddressLine1'],
                            'AddressLine2': row['AddressLine2'],
                            'City': row['City'],
                            'State': row['State'],
                            'Country': row['Country'],
                            'Status': row['Status'],
                            'OrderDate': row['OrderDate'],
                            'ProdReleaseDate': row['ProdReleaseDate'],
                            'SchedShipDate': row['SchedShipDate'],
                            'TargetArrival': row['TargetArrival'],
                            'NotBefore': row['NotBefore'],
                            'ShipVia': row['ShipVia'],
                            'LastModifiedOn': row['LastModifiedOn']
                        }
                    )
                    skus.append(sku)
            order_skus[order] = skus
        return order_skus

    def shrink_to_square(self, w, h, x, dim_to_shrink):
        """
        Shrinks the area of a rectangle by a multiplier `x`, changing only one dimension
        """
        if not (0 < x < 1):
            self.show_alert("Error", "Shrink multiplier must be between 0 and 1.", "error")

        original_area = w * h
        new_area = original_area * x

        if dim_to_shrink.lower() == 'height':
            # Option 2: change height, keep width
            new_h2 = new_area / w
            return (w, new_h2)
        elif dim_to_shrink.lower() == 'width':
            # Option 1: change width, keep height
            new_w1 = new_area / h
            return (new_w1, h)
        else:
            # shrink smaller dim if not specified
            if w < h:
                new_w1 = new_area / h
                return (new_w1, h)
            else:
                new_h2 = new_area / w
                return (w, new_h2)

    def write_optimized_bundles(self, workbook, order_bundles: dict):
        """
        Write the packed bundles to a new sheet in the workbook
        """
        # create a new sheet for the optimized bundles
        if "SO-PackExportData" in workbook.sheetnames:
            del workbook["SO-PackExportData"]
        if not self.append_data:
            if "Optimized_Bundles" in workbook.sheetnames:
                del workbook["Optimized_Bundles"]
            optimized_sheet = workbook.create_sheet("Optimized_Bundles")
        else:
            optimized_sheet = workbook["Optimized_Bundles"]

        # write headers
        intersect_headers = ['Can_be_bottom', 'Dim_shrink']
        # remove intersect headers from the main headers
        self.headers = [header for header in self.headers if header not in intersect_headers]
        # add headers
        self.headers.insert(6, 'TotalPcs')
        self.headers.insert(3, 'ApprovedBy')
        self.headers.insert(3, 'ReviewedBy')
        self.headers.insert(3, 'Machine')
        self.headers.insert(2, 'BundleNbr')
        self.headers.append('OptimizedOn')

        if self.set_unit == 'imperial':
            for i, header in enumerate(self.headers):
                if header == 'Width_mm':
                    self.headers[i] = 'Width_in'
                elif header == 'Height_mm':
                    self.headers[i] = 'Height_in'
                elif header == 'Length_mm':
                    self.headers[i] = 'Length_in'
                elif header == 'Weight_kg':
                    self.headers[i] = 'Weight_lbs'

            length_divisor = 25.4
            weight_multiplier = 2.20462
        else:
            length_divisor = 1
            weight_multiplier = 1

        if not self.append_data:
            optimized_sheet.append(self.headers)

        bottom_groups = []

        # add data for each order's bundles
        for orderIdx, order in enumerate(order_bundles.keys()):
            bundles = order_bundles[order]
            self.ui.progressBar.setValue(int(round(90 + 5 * ((orderIdx + 1) / len(order_bundles)))))
            QApplication.processEvents()

            # add a row with total order summary (only if there are bundles)
            if bundles:
                total_sub_bundles = sum([len(bundle.skus) for bundle in bundles])
                total_pcs = sum([sku.bundleqty for bundle in bundles for sku in bundle.skus])

                total_weight = sum([bundle.get_total_weight() for bundle in bundles])
                optimized_sheet.append([
                    bundles[0].skus[0].data['OrderType'],
                    order,
                    'ALL',  # BundleNbr
                    '',  # Bdl_Override
                    '',  # Machine
                    '',  # ReviewedBy
                    '',  # ApprovedBy
                    'Total_Order',
                    total_sub_bundles,
                    'N/A',
                    total_pcs,
                    'N/A',
                    'N/A',
                    'N/A',
                    round(total_weight * weight_multiplier),
                    '',
                    'Total Order Summary',
                    bundles[0].skus[0].data['ShipTo'],
                    bundles[0].skus[0].data['AddressLine1'],  # AddressLine1
                    bundles[0].skus[0].data['AddressLine2'],  # AddressLine2
                    bundles[0].skus[0].data['City'],  # City
                    bundles[0].skus[0].data['State'],  # State
                    bundles[0].skus[0].data['Country'],  # Country
                    bundles[0].skus[0].data['Status'],  # Status
                    bundles[0].skus[0].data['OrderDate'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['OrderDate'] else None,  # OrderDate
                    bundles[0].skus[0].data['ProdReleaseDate'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['ProdReleaseDate'] else None,  # ProdReleaseDate
                    bundles[0].skus[0].data['SchedShipDate'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['SchedShipDate'] else None,  # SchedShipDate
                    bundles[0].skus[0].data['TargetArrival'],  # TargetArrival
                    bundles[0].skus[0].data['NotBefore'],  # NotBefore
                    bundles[0].skus[0].data['ShipVia'],  # ShipVia
                    bundles[0].skus[0].data['LastModifiedOn'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['LastModifiedOn'] else None,  # LastModifiedOn
                    datetime.now().strftime("%Y-%m-%d"),
                ])

            # add missing/removed skus as part of bundle "0"
            if order in [sku.data['OrderNbr'] for sku in self.missingDataSKUs]:

                order_missing = [sku for sku in self.missingDataSKUs if sku.data['OrderNbr'] == order]
                # add a summary row for missing SKUs
                optimized_sheet.append([
                    order_missing[0].data['OrderType'],
                    order,
                    '0_ALL',  # BundleNbr
                    '',  # Bdl_Override
                    '',  # Machine
                    '',  # ReviewedBy
                    '',  # ApprovedBy
                    'Missing_SKUs',
                    len(order_missing),  # TotalPcs
                    'N/A',  # BundleQty
                    'N/A',  # Total Bundle Qty
                    'N/A',  # Width
                    'N/A',  # Height
                    'N/A',  # Length
                    'N/A',  # Weight
                    '',  # UOM
                    'Missing SKUs Summary',
                    order_missing[0].data['ShipTo'],
                    order_missing[0].data['AddressLine1'],  # AddressLine1
                    order_missing[0].data['AddressLine2'],  # AddressLine2
                    order_missing[0].data['City'],  # City
                    order_missing[0].data['State'],  # State
                    order_missing[0].data['Country'],  # Country
                    order_missing[0].data['Status'],  # Status
                    order_missing[0].data['OrderDate'].strftime("%Y-%m-%d") if order_missing[0].data['OrderDate'] else None,  # OrderDate
                    order_missing[0].data['ProdReleaseDate'].strftime("%Y-%m-%d") if order_missing[0].data['ProdReleaseDate'] else None,  # ProdReleaseDate
                    order_missing[0].data['SchedShipDate'].strftime("%Y-%m-%d") if order_missing[0].data['SchedShipDate'] else None,  # SchedShipDate
                    order_missing[0].data['TargetArrival'],  # TargetArrival
                    order_missing[0].data['NotBefore'],  # NotBefore
                    order_missing[0].data['ShipVia'],  # ShipVia
                    order_missing[0].data['LastModifiedOn'].strftime("%Y-%m-%d") if order_missing[0].data['LastModifiedOn'] else None,  # LastModifiedOn
                    datetime.now().strftime("%Y-%m-%d"),
                ])

                # add a row for each missing SKU
                written_skus = set()  # to avoid writing the same SKU multiple times
                for sku in order_missing:
                    if sku.id in written_skus:
                        continue
                    else:
                        written_skus.add(sku.id)
                        # count the number of identical SKUs in the order
                        order_skus = [s for s in self.missingDataSKUs if s.id == sku.id]
                        quantity = len(order_skus)
                    if sku.data['OrderNbr'] == order:
                        optimized_sheet.append([
                            sku.data['OrderType'],
                            order,
                            0,
                            sku.data['Bdl_Override'],
                            '',  # Machine
                            '',  # ReviewedBy
                            '',  # ApprovedBy
                            sku.id,
                            quantity,
                            round(sku.bundleqty) if sku.bundleqty else "N/A",  # default to 1 if bundleqty is None
                            "N/A" if not sku.bundleqty else round(quantity * sku.bundleqty),
                            sku.width / length_divisor if sku.width else "N/A",
                            sku.height / length_divisor if sku.height else "N/A",
                            sku.length / length_divisor if sku.length else "N/A",
                            round(sku.weight * weight_multiplier) if sku.weight else "N/A",
                            sku.data['UOM'],
                            sku.desc,
                            sku.data['ShipTo'],
                            sku.data['AddressLine1'],
                            sku.data['AddressLine2'],
                            sku.data['City'],
                            sku.data['State'],
                            sku.data['Country'],
                            sku.data['Status'],
                            sku.data['OrderDate'].strftime("%Y-%m-%d") if sku.data['OrderDate'] else None,
                            sku.data['ProdReleaseDate'].strftime("%Y-%m-%d") if sku.data['ProdReleaseDate'] else None,
                            sku.data['SchedShipDate'].strftime("%Y-%m-%d") if sku.data['SchedShipDate'] else None,
                            sku.data['TargetArrival'],
                            sku.data['NotBefore'],
                            sku.data['ShipVia'],
                            sku.data['LastModifiedOn'].strftime("%Y-%m-%d") if sku.data['LastModifiedOn'] else None,
                            datetime.now().strftime("%Y-%m-%d"),
                        ])

            order_row_count = 0
            for bundle_index, bundle in enumerate(bundles):
                # get quantity of each SKU in the bundle (including stacked quantities)
                sku_counts = {}
                for sku in bundle.skus:
                    if sku.id not in sku_counts:
                        sku_counts[sku.id] = {'qty': 0, 'sku': sku}
                    sku_counts[sku.id]['qty'] += 1

                # calculate bundle actual dimensions and weight
                actual_width, actual_height, _ = bundle.get_actual_dimensions()
                total_weight = bundle.get_total_weight()
                lumber = self.lumber_height if all([sku.rotated is False for sku in bundle.skus]) else 0

                # add summary row for the bundle
                optimized_sheet.append([
                    bundle.skus[0].data['OrderType'],
                    order,
                    f'{bundle_index + 1}_ALL',  # BundleNbr
                    bundle.skus[0].data['Bdl_Override'] if bundle.skus[0].data['Bdl_Override'] else '',  # Bdl_Override
                    bundle.packing_machine,  # Machine
                    '',  # ReviewedBy
                    '',  # ApprovedBy
                    f'Total_Bundle_{bundle_index + 1}',  # SKU
                    len(bundle.skus),  # TotalPcs
                    'N/A',  # BundleQty
                    sum(round(sku.bundleqty) for sku in bundle.skus),
                    round((actual_width + self.packaging_width) / length_divisor),
                    round((actual_height + self.packaging_height + lumber) / length_divisor),
                    round(bundle.max_length / length_divisor),
                    round(total_weight * weight_multiplier),
                    '',  # UOM
                    f'Bundle {bundle_index + 1} Summary',  # Description
                    bundle.skus[0].data['ShipTo'],
                    bundle.skus[0].data['AddressLine1'],  # AddressLine1
                    bundle.skus[0].data['AddressLine2'],  # AddressLine2
                    bundle.skus[0].data['City'],  # City
                    bundle.skus[0].data['State'],  # State
                    bundle.skus[0].data['Country'],  # Country
                    bundle.skus[0].data['Status'],  # Status
                    bundle.skus[0].data['OrderDate'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['OrderDate'] else None,  # OrderDate
                    bundle.skus[0].data['ProdReleaseDate'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['ProdReleaseDate'] else None,  # ProdReleaseDate
                    bundle.skus[0].data['SchedShipDate'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['SchedShipDate'] else None,  # SchedShipDate
                    bundle.skus[0].data['TargetArrival'],  # TargetArrival
                    bundle.skus[0].data['NotBefore'],  # NotBefore
                    bundle.skus[0].data['ShipVia'],  # ShipVia
                    bundle.skus[0].data['LastModifiedOn'].strftime("%Y-%m-%d") if bundles[0].skus[0].data['LastModifiedOn'] else None,  # LastModifiedOn
                    datetime.now().strftime("%Y-%m-%d"),
                ])

                packaging_skus_active = False
                packaging_idx = 2
                # write each SKU in the bundle to the sheet
                for sku_id, sku_data in sku_counts.items():
                    # fix length of SKU ID
                    if sku_data['sku'].length == 3650:
                        sku_data['sku'].length = 3680

                    if "Pack_Angle" in sku_id and not packaging_skus_active:
                        packaging_skus_active = True
                        packaging_idx = optimized_sheet.max_row + 2
                    # check if data is None (this happens for Packaging SKUs)
                    if sku_data['sku'].data is None:
                        # give data from another SKU in the order, since they are the same (except UOM)
                        for _, nested_sku_data in sku_counts.items():
                            if nested_sku_data['sku'].data is not None:
                                sku_data['sku'].data = nested_sku_data['sku'].data
                                sku_data['sku'].data['UOM'] = ''
                                break
                    try:
                        optimized_sheet.append([
                            sku_data['sku'].data['OrderType'],
                            order,
                            bundle_index + 1,
                            sku_data['sku'].data['Bdl_Override'],
                            bundle.packing_machine,
                            '',  # ReviewedBy
                            '',  # ApprovedBy
                            sku_id,
                            sku_data['qty'],
                            round(sku_data['sku'].bundleqty),
                            round(sku_data['qty'] * sku_data['sku'].bundleqty),
                            round(sku_data['sku'].width / length_divisor),
                            round(sku_data['sku'].height / length_divisor),
                            round(sku_data['sku'].length / length_divisor),
                            round(sku_data['sku'].weight * weight_multiplier),
                            sku_data['sku'].data['UOM'],
                            sku_data['sku'].desc,
                            sku_data['sku'].data['ShipTo'],
                            sku_data['sku'].data['AddressLine1'],
                            sku_data['sku'].data['AddressLine2'],
                            sku_data['sku'].data['City'],
                            sku_data['sku'].data['State'],
                            sku_data['sku'].data['Country'],
                            sku_data['sku'].data['Status'],
                            sku_data['sku'].data['OrderDate'].strftime("%Y-%m-%d") if sku_data['sku'].data['OrderDate'] else None,
                            sku_data['sku'].data['ProdReleaseDate'].strftime("%Y-%m-%d") if sku_data['sku'].data['ProdReleaseDate'] else None,
                            sku_data['sku'].data['SchedShipDate'].strftime("%Y-%m-%d") if sku_data['sku'].data['SchedShipDate'] else None,
                            sku_data['sku'].data['TargetArrival'],
                            sku_data['sku'].data['NotBefore'],
                            sku_data['sku'].data['ShipVia'],
                            sku_data['sku'].data['LastModifiedOn'].strftime("%Y-%m-%d") if sku_data['sku'].data['LastModifiedOn'] else None,
                            datetime.now().strftime("%Y-%m-%d"),
                        ])
                        order_row_count += 1
                    except Exception as e:
                        self.show_alert("Error", f"Error writing SKU {sku_id} to the sheet: {e}", "error")
                        return
                if packaging_skus_active:
                    bottom_groups.extend(range(packaging_idx - 1, optimized_sheet.max_row))
            # add a blank row after each order's bundles
            optimized_sheet.append([])

        # update text
        self.ui.progressLabel.setText("Saving Excel file...")
        QApplication.processEvents()

        # create a table over the data
        if "OptimizedBundlesTable" in optimized_sheet.tables:
            del optimized_sheet.tables["OptimizedBundlesTable"]
        table = openpyxl.worksheet.table.Table(displayName="OptimizedBundlesTable", ref=optimized_sheet.dimensions, tableStyleInfo=openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True))
        # resize the table to fit the data
        optimized_sheet.add_table(table)

        optimized_sheet.row_dimensions[optimized_sheet.max_row+1].outlineLevel = 1

        # if two blank rows in a row, remove the first one
        for row in range(optimized_sheet.max_row, 1, -1):
            # 2 consecutive blank rows
            if optimized_sheet[row][0].value is None and optimized_sheet[row - 1][0].value is None:
                optimized_sheet.delete_rows(row - 1, 1)
                optimized_sheet.row_dimensions[row + 1].outlineLevel = 1
            if optimized_sheet.row_dimensions[row].outlineLevel == 3:
                continue
            # if the row is blank, set the outline level to 0
            if row in bottom_groups:
                optimized_sheet.row_dimensions[row].outlineLevel = 3
            elif type(optimized_sheet[row][2].value) is int:
                optimized_sheet.row_dimensions[row].outlineLevel = 2
            elif optimized_sheet[row][0].value is None:
                optimized_sheet.row_dimensions[row].outlineLevel = 0
                optimized_sheet.row_dimensions[row + 1].outlineLevel = 0
            # if the row is not blank, set the outline level to 1
            else:
                optimized_sheet.row_dimensions[row].outlineLevel = 1
        optimized_sheet.row_dimensions[2].outlineLevel = 0  # first row is the header, no grouping

        for row in range(1, optimized_sheet.max_row + 1):
            optimized_sheet.row_dimensions[row].hidden = False

        # save the workbook
        try:
            if self.append_data:
                workbook.save(self.ui.appendDir.text())
            else:
                workbook.save(f"{self.workingDir}/Optimized_Bundles.xlsx")
        except Exception as e:
            self.show_alert("Error", f"Error saving the file. Is it already open? Error: {e}", "error")
            return

    def show_alert(self, title, message, type="warning") -> None:
        """
        Show an alert dialog with the given title and message
        """
        msg_box = QMessageBox(self.Widget)
        if type == "help":
            msg_box.setTextFormat(QtCore.Qt.TextFormat.RichText)
            type = "info"
        msg_box.setWindowTitle(title)
        msg_box.setText(message)

        # Set the icon based on type
        if type == "warning":
            msg_box.setIcon(QMessageBox.Icon.Warning)
        elif type == "info":
            msg_box.setIcon(QMessageBox.Icon.Information)
        elif type == "error":
            msg_box.setIcon(QMessageBox.Icon.Critical)

        # Set stylesheet
        msg_box.setStyleSheet("""
            QMessageBox {
                color: white;
            }
            QLabel {
                color: white;
            }
            QLabel a {
                color: #2980b9;
            }
            QPushButton {
                background-color: #eaeaea;
                color: #090909;
            }
        """)


        msg_box.exec()
