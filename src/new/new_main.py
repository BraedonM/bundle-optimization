from PyQt6.QtWidgets import QMessageBox, QWidget
from BundleQtGui import Ui_BundleOptimizer

import openpyxl
import pandas as pd
from tkinter import filedialog
import os
from math import ceil

from bundle_classes import SKU
from bundle_visualize import visualize_bundles
from bundle_packing import pack_skus

class Ui_MainWindow:
    def __init__(self):
        self.Widget = QWidget()
        self.ui = Ui_BundleOptimizer()
        self.ui.setupUi(self.Widget)

        self.setupUi()
        self.Widget.show()

    def setupUi(self):
        # Initialize values and connect signals (camelCase, while helper methods are snake_case)
        self.ui.fileBrowse.clicked.connect(self.getInputWorkbook)
        self.ui.openExample.clicked.connect(self.openExampleFile)
        self.ui.optimizeBundles.clicked.connect(self.optimizeBundles)
        self.ui.openImages.clicked.connect(self.openImages)
        self.ui.openExcel.clicked.connect(self.openExcel)
        self.ui.helpButton.clicked.connect(self.openHelp)

        self.data = pd.DataFrame()
        self.maxWidth = 590
        self.maxHeight = 590
        self.maxLength = 3680
        self.workingDir = None  # to hold the directory of the selected Excel file
        self.missingDataSKUs = []  # to hold SKUs that are missing data in the Excel file

## QT Connection Methods (camelCase)

    def getInputWorkbook(self):
        """
        Open a dialog to pick an Excel file, return the workbook object
        """
        # get excel file path from user
        path = filedialog.askopenfilename(
            title="Select an Excel file",
            filetypes=[("Excel files", "*.xlsx; *.xls; *.xlsm")],
        )
        if not path:
            return None

        self.workingDir = path.rsplit('/', 1)[0]  # get the directory of the selected file

        # load the workbook and read the sheet "SO_Input"
        self.workbook = openpyxl.load_workbook(path)
        self.ui.excelDir.setText(path)

    def openExampleFile(self):
        """
        Open the example Excel file to the user, in Excel
        """
        example_path = os.path.join(os.path.dirname(__file__), 'SO_Input_Example.xlsx')
        if os.path.exists(example_path):
            os.startfile(example_path)
        else:
            self.show_alert("File Not Found", "Example file not found.")

    def optimizeBundles(self):
        """
        Optimize bundles based on the data from the Excel file
        """
        self.ui.progressLabel.setText("Getting data...")
        self.ui.progressBar.setValue(10)
        data = self.get_data(self.workbook)
        if data.empty:
            self.show_alert("Error", "Invalid path for Excel file.", "error")
            return

        # get unique orders
        unique_orders = data['OrderNbr'].unique()

        # get array of rows in sorted_data that belong to each order
        order_rows = {order: data[data['OrderNbr'] == order] for order in unique_orders}

        # create SKU objects for each order
        order_skus = self.create_sku_objects(order_rows)

        order_skus = self.remove_invalids(order_skus)

        images_dir = f"{self.workingDir}/images"
        os.makedirs(images_dir, exist_ok=True)

        images_dir = f"{self.workingDir}/images"
        os.makedirs(images_dir, exist_ok=True)

        # pack each order's SKUs into bundles
        order_bundles = {}
        for order, skus in order_skus.items():
            if skus == []:
                continue
            self.ui.progressBar.setValue(int(round(10 + 80 * (list(order_skus.keys()).index(order) + 1) / len(order_skus))))
            self.ui.progressLabel.setText(f"Packing order {order}...")
            bundles = pack_skus(skus, self.maxWidth, self.maxHeight)
            order_bundles[order] = bundles

            visualize_bundles(bundles, f"{images_dir}/Order_{order}.png")
        self.images_dir = images_dir

        # write the packed bundles to a new sheet in the workbook
        self.write_optimized_bundles(self.workbook, order_bundles)

        self.ui.progressBar.setValue(100)
        self.ui.progressLabel.setText("Packing complete!")

        if self.missingDataSKUs:
            self.show_alert("Missing Data", "There exist InventoryIDs that are missing data in the Excel file\nand have been excluded from optimization.\n\nPlease check the 'missing_data_skus.txt' file for details.", "warning")
            # write the missing SKUs to a file
            with open(f"{self.workingDir}/missing_data_skus.txt", 'w') as f:
                f.write("The following SKUs are missing data and could not be included in the optimization:\n\n")
                f.write('\n'.join(self.missingDataSKUs))

    def openImages(self):
        """
        Open the images directory in the file explorer
        """
        if hasattr(self, 'images_dir'):
            os.startfile(self.images_dir)
        else:
            self.show_alert("No Images", "No images have been generated yet. Please optimize bundles first.", "error")

    def openExcel(self):
        """
        Open the optimized bundles Excel file in the file explorer
        """
        if not self.workingDir:
            self.show_alert("Error", "No file found", "error")
            return
        excel_path = f"{self.workingDir}/Optimized_Bundles.xlsx"
        if os.path.exists(excel_path):
            os.startfile(excel_path)
        else:
            self.show_alert("Error", "Optimized bundles file not found. Please optimize bundles first.", "error")

    def openHelp(self):
        """
        Open the help file in the file explorer
        """
        self.show_alert("Help", """
        For additional support, please contact Aionex Solutions Ltd.
        (604-309-8975)

        To use this tool, follow these steps:

        1. Click on 'Browse' to select your Excel file containing SKU data.
            - If you are unsure how to format your Excel file,
              you can click on 'Open Example File' to open a sample file.
            - Please format your Excel file according to the example provided.

        2. Click on 'Perform Bundle Optimization' to start the
           optimization process.
           Wait for the progress bar to complete.

        3. Once the optimization is complete, you can:
            - Click on 'Open Images Folder' to view the generated bundle images.
            - Click on 'Open Resultant Excel File' to view the optimized
              bundle data in an Excel file.

        NOTE: All files are generated in the same directory as the input Excel file.
        """, type="info")

## Helper Methods (snake_case)

    def get_sub_bundle_data_sheet(self):
        """
        Open a dialog to pick an Excel file, return the sheet "Sub-Bundle_Data"
        """
        # load the workbook and read the sheet "Sub-Bundle_Data"
        path = os.path.join(os.path.dirname(__file__), 'Sub-Bundle_Data.xlsx')
        workbook = openpyxl.load_workbook(path)
        sheet = workbook["Sub-Bundle_Data"]
        if not sheet:
            self.show_alert("Warning", "Sheet 'Sub-Bundle_Data' not found in file.")
            return None
        return sheet

    def get_data(self, workbook):
        """
        Read data from the 'SO_Input' sheet of the workbook
        """
        # get the "SO_PackExportData" sheet
        so_input = workbook["SO-PackExportData"]
        if not so_input:
            self.show_alert("Warning", "Sheet 'SO-PackExportData' is empty or not found. Using first sheet in the file instead.")
            so_input = workbook.active  # fallback to the first sheet if is not found
        sb_data = self.get_sub_bundle_data_sheet()

        df = pd.DataFrame(columns=[cell.value for cell in so_input[1]])
        sb_df = pd.DataFrame(columns=[cell.value for cell in sb_data[2]])
        # read all rows from the sheets
        for row in so_input.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            df.loc[len(df)] = row
        for row in sb_data.iter_rows(min_row=3, values_only=True):
            # check for any empty cells in the row
            if any(cell is None for cell in [el for el in row][2:]):
                continue
            sb_df.loc[len(sb_df)] = row

        # remove any rows with SKUs that are in the missingDataSKUs list
        df = df[~df['InventoryID'].isin(self.missingDataSKUs)]

        # check if the required columns are present
        self.headers = ["OrderType", "OrderNbr", "Bdl_Override", "InventoryID", "Quantity", "Pcs/Bundle", "Can_be_bottom",
                   "Width_mm", "Height_mm", "Length_mm", "Weight_kg", "UOM", "Description",
                   "ShipTo", "AddressLine1", "AddressLine2", "City", "State", "Country", "Status", "OrderDate",
                   "ProdReleaseDate", "SchedShipDate", "TargetArrival", "NotBefore", "ShipVia", "LastModifiedOn"]
        for col in self.headers:
            if col not in df.columns:
                # add the column with default values
                df[col] = None

        # put data from sb_df into df rows
        for _, row in sb_df.iterrows():
            sku_id = row['SKU']
            df_rows = []
            if row['Bottom Row Acceptable']:
                can_be_bottom = True
            else:
                can_be_bottom = False
            for index, df_row in df.iterrows():
                try:
                    if sku_id in df_row['InventoryID']:
                        df_rows.append(index)
                except TypeError:
                    # if df_row['InventoryID'] is None, skip this row
                    continue
            for df_row_idx in df_rows:
                # update the Pcs/Bundle column with the value from sb_df
                df.loc[df_row_idx, 'Pcs/Bundle'] = row['Qty/bundle']
                df.loc[df_row_idx, 'Width_mm'] = row['Width (mm)']
                df.loc[df_row_idx, 'Height_mm'] = row['Height (mm)']
                df.loc[df_row_idx, 'Length_mm'] = row['Length (mm)']
                df.loc[df_row_idx, 'Weight_kg'] = row['Weight kg/length']
                df.loc[df_row_idx, 'Can_be_bottom'] = can_be_bottom

        # iterate through df and convert qty (in pieces) to qty (in bundles)
        for index, row in df.iterrows():
            if row['Pcs/Bundle'] is not None and row['Quantity'] is not None:
                try:
                    # convert Quantity to Pcs/Bundle
                    df.loc[index, 'Quantity'] = ceil(row['Quantity'] / row['Pcs/Bundle'])
                except ZeroDivisionError:
                    self.show_alert("Error", f"Pcs/Bundle cannot be zero for SKU {row['InventoryID']}. Please check the input data.", "error")
                    return pd.DataFrame()

        return df

    def remove_invalids(self, order_skus: dict) -> dict:
        """
        Iterate through the order_skus dictionary and remove any SKUs that have None as width, height, length, or weight.
        """
        for order, skus in order_skus.items():
            valid_skus = []
            for sku in skus:
                if sku.width is None or sku.height is None or sku.length is None or sku.weight is None:
                    if sku.id not in self.missingDataSKUs:
                        self.missingDataSKUs.append(sku.id)  # add to missing data SKUs list
                else:
                    valid_skus.append(sku)
            order_skus[order] = valid_skus
        return order_skus

    def create_sku_objects(self, order_rows: dict):
        """
        Create arrays of SKU objects for each order
        """
        order_skus = {}
        for order, rows in order_rows.items():
            skus = []
            for _, row in rows.iterrows():
                # get quantity of SKU from the row
                quantity = row['Quantity']
                invID = row['InventoryID'].strip()
                newLength = row['Length_mm']
                if newLength is not None:
                    if 3600 <= newLength <= 3700:
                        newLength = 3650
                for _ in range(int(abs(ceil(quantity)))):
                    sku = SKU(
                        id=invID,
                        bundleqty=row['Pcs/Bundle'],
                        width=row['Width_mm'],
                        height=row['Height_mm'],
                        length=newLength,
                        weight=row['Weight_kg'],
                        desc=row['Description'],
                        can_be_bottom=row['Can_be_bottom'],
                        data={
                            'OrderType': row['OrderType'],
                            'OrderNbr': row['OrderNbr'],
                            'UOM': row['UOM'],
                            'Bdl_Override': row['Bdl_Override'] if pd.notna(row['Bdl_Override']) else None,
                            'ShipTo': row['ShipTo'],
                            'AddressLine1': row['AddressLine1'],
                            'AddressLine2': row['AddressLine2'],
                            'City': row['City'],
                            'State': row['State'],
                            'Country': row['Country'],
                            'Status': row['Status'],
                            'OrderDate': row['OrderDate'],
                            'ProdReleaseDate': row['ProdReleaseDate'],
                            'SchedShipDate': row['SchedShipDate'],
                            'TargetArrival': row['TargetArrival'],
                            'NotBefore': row['NotBefore'],
                            'ShipVia': row['ShipVia'],
                            'LastModifiedOn': row['LastModifiedOn']
                        }
                    )
                    skus.append(sku)
            order_skus[order] = skus
        return order_skus

    def write_optimized_bundles(self, workbook, order_bundles: dict):
        """
        Write the packed bundles to a new sheet in the workbook
        """
        # create a new sheet for the optimized bundles
        if "Optimized_Bundles" in workbook.sheetnames:
            del workbook["Optimized_Bundles"]
        optimized_sheet = workbook.create_sheet("Optimized_Bundles")

        # write headers
        intersect_headers = ['Can_be_bottom']
        # remove intersect headers from the main headers
        self.headers = [header for header in self.headers if header not in intersect_headers]
        # add headers
        self.headers.insert(6, 'TotalPcs')
        self.headers.insert(2, 'BundleNbr')
        optimized_sheet.append(self.headers)

        top_groups = []
        bottom_groups = []

        # add data for each order's bundles
        for order, bundles in order_bundles.items():
            order_row_count = 0
            for bundle_index, bundle in enumerate(bundles):
                # get quantity of each SKU in the bundle (including stacked quantities)
                sku_counts = {}
                for sku in bundle.skus:
                    if sku.id not in sku_counts:
                        sku_counts[sku.id] = {'qty': 0, 'sku': sku}
                    sku_counts[sku.id]['qty'] += sku.stacked_quantity

                # calculate bundle actual dimensions and weight
                actual_width, actual_height, _ = bundle.get_actual_dimensions()
                total_weight = bundle.get_total_weight()

                packaging_skus_active = False
                packaging_idx = 0
                # write each SKU in the bundle to the sheet
                for sku_id, sku_data in sku_counts.items():
                    if "Pack_Angle" in sku_id and not packaging_skus_active:
                        packaging_skus_active = True
                        packaging_idx = optimized_sheet.max_row + 2
                    # check if data is None (this happens for Packaging SKUs)
                    if sku_data['sku'].data is None:
                        # give data from another SKU in the order, since they are the same
                        for _, nested_sku_data in sku_counts.items():
                            if nested_sku_data['sku'].data is not None:
                                sku_data['sku'].data = nested_sku_data['sku'].data
                                break
                    try:
                        optimized_sheet.append([
                            sku_data['sku'].data['OrderType'],
                            order,
                            bundle_index + 1,
                            sku_data['sku'].data['Bdl_Override'],
                            sku_id,
                            sku_data['qty'],
                            sku_data['sku'].bundleqty,
                            sku_data['qty'] * sku_data['sku'].bundleqty,
                            actual_width,
                            actual_height,
                            bundle.max_length,
                            total_weight,
                            sku_data['sku'].data['UOM'],
                            sku_data['sku'].desc,
                            sku_data['sku'].data['ShipTo'],
                            sku_data['sku'].data['AddressLine1'],
                            sku_data['sku'].data['AddressLine2'],
                            sku_data['sku'].data['City'],
                            sku_data['sku'].data['State'],
                            sku_data['sku'].data['Country'],
                            sku_data['sku'].data['Status'],
                            sku_data['sku'].data['OrderDate'],
                            sku_data['sku'].data['ProdReleaseDate'],
                            sku_data['sku'].data['SchedShipDate'],
                            sku_data['sku'].data['TargetArrival'],
                            sku_data['sku'].data['NotBefore'],
                            sku_data['sku'].data['ShipVia'],
                            sku_data['sku'].data['LastModifiedOn'],
                        ])
                        order_row_count += 1
                    except Exception as e:
                        self.show_alert("Error", f"Error writing SKU {sku_id} to the sheet: {e}", "error")
                        return
                bottom_groups.append([packaging_idx, optimized_sheet.max_row])
            # group the rows by order number in Excel
            top_groups.append([optimized_sheet.max_row - order_row_count + 2, optimized_sheet.max_row])
            # add a blank row after each order's bundles
            optimized_sheet.append([])

        # create a table over the data
        table = openpyxl.worksheet.table.Table(displayName="OptimizedBundlesTable", ref=optimized_sheet.dimensions, tableStyleInfo=openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True))
        # resize the table to fit the data
        optimized_sheet.add_table(table)

        for group in top_groups:
            # group orders together
            optimized_sheet.row_dimensions.group(start=group[0], end=group[1], hidden=True)
            for row in range(group[0], group[1]):
                optimized_sheet.row_dimensions[row].outlineLevel = 1

        for group in bottom_groups:
            # group packaging SKUs together
            optimized_sheet.row_dimensions.group(start=group[0], end=group[1], hidden=True)
            for row in range(group[0], group[1]):
                optimized_sheet.row_dimensions[row].outlineLevel = 2
        # save the workbook
        try:
            workbook.save(f"{self.workingDir}/Optimized_Bundles.xlsx")
        except Exception as e:
            self.show_alert("Error", f"Error saving the file. Is it already open? Error: {e}", "error")
            return

    def show_alert(self, title, message, type="warning") -> None:
        """
        Show an alert dialog with the given title and message
        """
        if type == "warning":
            QMessageBox.warning(self.Widget, title, message)
        elif type == "info":
            QMessageBox.information(self.Widget, title, message)
        elif type == "error":
            QMessageBox.critical(self.Widget, title, message)
