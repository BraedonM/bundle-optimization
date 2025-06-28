import openpyxl
import pandas as pd
from tkinter import filedialog
from typing import List
import os
from itertools import groupby

from bundle_classes import SKU, Bundle
from bundle_visualize import visualize_bundles

global maxWidth
global maxHeight
global workingDir

maxWidth = 590
maxHeight = 590

def get_workbook():
    """
    Open a dialog to pick an Excel file, return the workbook object
    """
    global workingDir
    # get excel file path from user
    path = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx; *.xls; *.xlsm")],
    )
    if not path:
        print("No workbook selected.")
        return None

    workingDir = path.rsplit('/', 1)[0]  # get the directory of the selected file

    # load the workbook and read the sheet "SO_Input"
    workbook = openpyxl.load_workbook(path)
    return workbook

def get_data(workbook):
    """
    Read data from the 'SO_Input' sheet of the workbook
    """
    # get the "SO_Input" sheet
    if "SO_Input" not in workbook.sheetnames:
        print("Sheet 'SO_Input' not found in the selected file.")
        return None
    sheet = workbook["SO_Input"]

    # create headers with dataframe
    headers = [cell.value for cell in sheet[1]] # first row as headers
    df = pd.DataFrame(columns=headers)
    # read all rows from the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        df.loc[len(df)] = row  # append each row to the dataframe

    return df

def pack_skus_with_pattern(skus: List[SKU], bundle_width: int, bundle_height: int) -> List[Bundle]:
    """
    Packs SKUs into bundles using a pattern-based algorithm:
    - First row: vertical orientation (shortest side as width)
    - Subsequent rows: horizontal until height matches vertical row
    - Then alternate back to vertical, and so on
    - Final greedy optimization to fill remaining space
    """

    def get_sku_dimensions(sku: SKU, vertical: bool):
        """Return (width, height) for SKU based on orientation"""
        if vertical:
            # Vertical: shortest side as width, longest as height
            return min(sku.width, sku.height), max(sku.width, sku.height)
        else:
            # Horizontal: longest side as width, shortest as height
            return max(sku.width, sku.height), min(sku.width, sku.height)

    def can_fit_in_bundle(sku: SKU, x: int, y: int, vertical: bool, bundle: Bundle) -> bool:
        """Check if SKU can fit at position with given orientation"""
        width, height = get_sku_dimensions(sku, vertical)

        # Check bundle boundaries
        if x + width > bundle.width or y + height > bundle.height:
            return False

        # Check for overlap with existing SKUs
        for placed_sku in bundle.skus:
            if (x < placed_sku.x + placed_sku.width and
                x + width > placed_sku.x and
                y < placed_sku.y + placed_sku.height and
                y + height > placed_sku.y):
                return False
        return True

    def can_fit_any_orientation(sku: SKU, x: int, y: int, bundle: Bundle) -> tuple:
        """Check if SKU can fit at position in any orientation. Returns (can_fit, rotated)"""
        # Try original orientation
        if (x + sku.width <= bundle.width and y + sku.height <= bundle.height):
            overlap = False
            for placed_sku in bundle.skus:
                if (x < placed_sku.x + placed_sku.width and
                    x + sku.width > placed_sku.x and
                    y < placed_sku.y + placed_sku.height and
                    y + sku.height > placed_sku.y):
                    overlap = True
                    break
            if not overlap:
                return True, False

        # Try rotated orientation
        if (x + sku.height <= bundle.width and y + sku.width <= bundle.height):
            overlap = False
            for placed_sku in bundle.skus:
                if (x < placed_sku.x + placed_sku.width and
                    x + sku.height > placed_sku.x and
                    y < placed_sku.y + placed_sku.height and
                    y + sku.width > placed_sku.y):
                    overlap = True
                    break
            if not overlap:
                return True, True

        return False, False

    def greedy_fill_remaining_space(bundle: Bundle, remaining_skus: List[SKU]) -> List[SKU]:
        """Fill remaining space in bundle using greedy algorithm"""
        still_remaining = remaining_skus.copy()

        # Generate candidate points from existing SKUs
        candidate_points = {(0, 0)}
        for placed_sku in bundle.skus:
            candidate_points.add((placed_sku.x + placed_sku.width, placed_sku.y))
            candidate_points.add((placed_sku.x, placed_sku.y + placed_sku.height))

        # Sort candidate points by y first, then x (bottom-left preference)
        sorted_points = sorted(candidate_points, key=lambda p: (p[1], p[0]))

        placed_any = True
        while placed_any and still_remaining:
            placed_any = False

            for point in sorted_points:
                x, y = point
                best_sku_idx = None
                best_rotation = False

                # Find the best SKU for this position (prefer heavier SKUs)
                for i, sku in enumerate(still_remaining):
                    can_fit, rotated = can_fit_any_orientation(sku, x, y, bundle)
                    if can_fit:
                        if best_sku_idx is None or sku.weight > still_remaining[best_sku_idx].weight:
                            best_sku_idx = i
                            best_rotation = rotated

                if best_sku_idx is not None:
                    # Place the best SKU
                    sku = still_remaining.pop(best_sku_idx)

                    if best_rotation:
                        sku.width, sku.height = sku.height, sku.width

                    bundle.add_sku(sku, x, y, best_rotation)

                    # Update candidate points
                    candidate_points.add((x + sku.width, y))
                    candidate_points.add((x, y + sku.height))
                    sorted_points = sorted(candidate_points, key=lambda p: (p[1], p[0]))

                    placed_any = True
                    break  # Start over with new candidate points

        return still_remaining

    def pack_bundle_with_pattern(available_skus: List[SKU], bundle: Bundle) -> List[SKU]:
        """Pack a single bundle following the vertical/horizontal pattern"""
        remaining_skus = available_skus.copy()
        current_y = 0
        is_vertical_row = True
        vertical_row_height = 0
        horizontal_start_y = 0

        while remaining_skus and current_y < bundle.height:
            # Find SKUs that can fit in current row orientation
            row_skus = []
            row_height = 0
            current_x = 0

            # Sort SKUs by weight (heaviest first) for each row
            remaining_skus.sort(key=lambda x: x.weight, reverse=True)

            for i, sku in enumerate(remaining_skus):
                width, height = get_sku_dimensions(sku, is_vertical_row)

                # Check if SKU fits in current row
                if (current_x + width <= bundle.width and
                    current_y + height <= bundle.height and
                    can_fit_in_bundle(sku, current_x, current_y, is_vertical_row, bundle)):

                    row_skus.append((i, sku, current_x, current_y, width, height))
                    current_x += width
                    row_height = max(row_height, height)

            if not row_skus:
                break  # No more SKUs fit

            # Place the SKUs in this row
            placed_in_row = []
            for i, sku, x, y, width, height in row_skus:
                # Determine if rotation is needed
                original_width, original_height = sku.width, sku.height
                rotated = False

                if is_vertical_row:
                    # Vertical row: want min dimension as width
                    if original_width > original_height:
                        rotated = True
                        sku.width, sku.height = sku.height, sku.width
                else:
                    # Horizontal row: want max dimension as width
                    if original_width < original_height:
                        rotated = True
                        sku.width, sku.height = sku.height, sku.width

                bundle.add_sku(sku, x, y, rotated)
                placed_in_row.append(i)

            # Remove placed SKUs from remaining list (in reverse order to maintain indices)
            for i in sorted(placed_in_row, reverse=True):
                remaining_skus.pop(i)

            # Move to next row
            current_y += row_height

            # Determine orientation for next row
            if is_vertical_row:
                # After placing a vertical row, switch to horizontal and record the height
                is_vertical_row = False
                vertical_row_height = row_height
                horizontal_start_y = current_y  # Start tracking horizontal section from current position
            else:
                # We're in horizontal mode - check if we should switch back to vertical
                # Calculate total height of horizontal section so far
                horizontal_section_height = current_y - horizontal_start_y

                # Only switch back to vertical when horizontal section matches vertical row height
                if horizontal_section_height >= vertical_row_height:
                    is_vertical_row = True
                    # Reset for next cycle
                    vertical_row_height = 0
                    horizontal_start_y = 0

        return remaining_skus

    # Main packing logic
    skus.sort(key=lambda x: x.weight, reverse=True)  # Sort by weight initially
    bundles: List[Bundle] = []
    remaining_skus = skus.copy()

    while remaining_skus:
        new_bundle = Bundle(bundle_width, bundle_height)

        # Check if any remaining SKU can fit in a bundle at all
        can_fit_any = False
        for sku in remaining_skus:
            vert_w, vert_h = get_sku_dimensions(sku, True)
            horiz_w, horiz_h = get_sku_dimensions(sku, False)

            if ((vert_w <= bundle_width and vert_h <= bundle_height) or
                (horiz_w <= bundle_width and horiz_h <= bundle_height)):
                can_fit_any = True
                break

        if not can_fit_any:
            break

        # Pack this bundle with pattern-based algorithm
        before_count = len(remaining_skus)
        remaining_skus = pack_bundle_with_pattern(remaining_skus, new_bundle)
        after_count = len(remaining_skus)

        # GREEDY OPTIMIZATION: Try to fill remaining space with leftover SKUs
        if remaining_skus and new_bundle.skus:
            remaining_skus = greedy_fill_remaining_space(new_bundle, remaining_skus)

        if before_count == after_count:
            # No progress made, skip the largest remaining SKU
            if remaining_skus:
                largest_sku = max(remaining_skus, key=lambda x: x.width * x.height)
                remaining_skus.remove(largest_sku)

        if new_bundle.skus:  # Only add bundles that have SKUs
            bundles.append(new_bundle)

    return bundles


# Replace the pack_skus function in main.py with this version
def pack_skus(skus: List[SKU], bundle_width: int, bundle_height: int) -> List[Bundle]:
    """
    Wrapper function to maintain compatibility with existing code
    """
    return pack_skus_with_pattern(skus, bundle_width, bundle_height)

def create_sku_objects(order_rows: dict):
    """
    Create arrays of SKU objects for each order
    """
    order_skus = {}
    for order, rows in order_rows.items():
        skus = []
        for _, row in rows.iterrows():
            # get quantity of SKU from the row
            quantity = row['SKU Qty']
            for _ in range(quantity):
                sku = SKU(
                    id=row['SKU'],
                    width=row['SKU Width (mm)'],
                    height=row['SKU Height (mm)'],
                    length=row['SKU Length (mm)'],
                    weight=row['SKU Weight (kg)'],
                    desc=row['SKU Description']
                )
                skus.append(sku)
        order_skus[order] = skus
    return order_skus

def write_optimized_bundles(workbook, order_bundles: dict):
    """
    Write the packed bundles to a new sheet in the workbook
    """
    # create a new sheet for the optimized bundles
    if "Optimized_Bundles" in workbook.sheetnames:
        del workbook["Optimized_Bundles"]
    optimized_sheet = workbook.create_sheet("Optimized_Bundles")

    # write headers
    headers = [
               "Order ID",
               "Bundle No.",
               "SKU",
               "Qty",
               "SKU Description",
               "Bundle Total Width (mm)",
               "Bundle Total Height (mm)",
               "Bundle Total Weight (kg)",
               "Note"
               ]
    optimized_sheet.append(headers)

    # add data for each order's bundles
    for order, bundles in order_bundles.items():
        for bundle_index, bundle in enumerate(bundles):
            # get quantity of each SKU in the bundle
            sku_counts = {}
            for sku in bundle.skus:
                if sku.id not in sku_counts:
                    sku_counts[sku.id] = {'qty': 0, 'description': sku.desc, 'weight': sku.weight}
                sku_counts[sku.id]['qty'] += 1
            # calculate bundle total width, height, and weight
            total_width = bundle.width
            total_height = bundle.height
            total_weight = sum(sku.weight for sku in bundle.skus)
            # write each SKU in the bundle to the sheet
            for sku_id, sku_data in sku_counts.items():
                optimized_sheet.append([
                    order,
                    bundle_index + 1,
                    sku_id,
                    sku_data['qty'],
                    sku_data['description'],
                    total_width,
                    total_height,
                    total_weight,
                    ""
                ])
        # add a blank row after each order's bundles
        optimized_sheet.append([])

    # create a table over the data
    table = openpyxl.worksheet.table.Table(displayName="OptimizedBundlesTable", ref=optimized_sheet.dimensions, tableStyleInfo=openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True))
    # resize the table to fit the data
    optimized_sheet.add_table(table)

    # save the workbook
    try:
        workbook.save(f"{workingDir}/Optimized_Bundles.xlsx")
    except Exception as e:
        print(f"Error saving the file. Is it already open? Error: {e}")
        return

if __name__ == "__main__":
    # get the workbook and data
    print("Please select the Excel file with order data")
    wb = get_workbook()
    if not wb:
        exit()
    data = get_data(wb)

    print("\nData loaded successfully! Sorting and packing SKUs...\n")

    # get unique orders
    unique_orders = data['Order ID'].unique()

    # get array of rows in sorted_data that belong to each order
    order_rows = {order: data[data['Order ID'] == order] for order in unique_orders}

    # create SKU objects for each order
    order_skus = create_sku_objects(order_rows)

    print("Saving images...")
    images_dir = f"{workingDir}/images"
    os.makedirs(images_dir, exist_ok=True)

    # pack each order's SKUs into bundles
    order_bundles = {}
    for order, skus in order_skus.items():
        bundles = pack_skus(skus, maxWidth, maxHeight)
        order_bundles[order] = bundles
        for i, bundle in enumerate(bundles):
            for sku in bundle.skus:
                rot = " (rotated)" if sku.rotated else ""
        visualize_bundles(bundles, f"{images_dir}/Order_{order}.png")

    print(f"\nImages saved in {images_dir}")

    # write the packed bundles to a new sheet in the workbook
    write_optimized_bundles(wb, order_bundles)

    print(f"Optimized bundles written to {workingDir}/Optimized_Bundles.xlsx\n\nProgram complete.")
